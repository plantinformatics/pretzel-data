# from
# https://chatgpt.com/share/67b5b1ca-c6c4-800e-bb28-3f7aa8dd4157
# https://chatgpt.com/canvas/shared/67b5ab81f6e08191962282b1d9d6c5e3

import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

def format_excel(output_file, df):
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Identify columns between 'alt' and 'index' (exclusive), excluding 'TYPE'
    if 'alt' in df.columns and 'index' in df.columns:
        start_col = df.columns.get_loc('alt') + 1
        end_col = df.columns.get_loc('index')
        columns_to_rotate = [df.columns[i] for i in range(start_col, end_col) if df.columns[i] != 'TYPE']
    else:
        columns_to_rotate = []
    
    # Formatting rules
    rotation_angle = 90
    narrow_column_width = 4
    color_mapping = {
        0: "B1C1E8",
        2: "FFC8AE"
    }
    
    # Apply formatting
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        if col_name in columns_to_rotate:
            cell.alignment = Alignment(textRotation=rotation_angle, horizontal="center", vertical="center")
            ws.column_dimensions[cell.column_letter].width = narrow_column_width
    
    # Apply cell colors based on values in specified columns
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=start_col + 1, max_col=end_col), start=2):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value in color_mapping:
                cell.fill = PatternFill(start_color=color_mapping[cell.value], end_color=color_mapping[cell.value], fill_type="solid")
    
    wb.save(output_file)

def main():
    if len(sys.argv) != 3:
        print("Usage: python script.py <input_CSV_file_name> <output_XLSX_file_name>")
        sys.exit(1)
    
    input_csv = sys.argv[1]
    output_xlsx = sys.argv[2]
    
    # Read CSV file
    df = pd.read_csv(input_csv)
    
    # Write to Excel
    df.to_excel(output_xlsx, index=False, engine='openpyxl')
    
    # Apply formatting
    format_excel(output_xlsx, df)
    
    print(f"Formatted Excel file saved as {output_xlsx}")

if __name__ == "__main__":
    main()
